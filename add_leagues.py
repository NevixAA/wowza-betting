"""
Add new leagues to the system by downloading from football-data.co.uk (mmz4281 format).

Leagues added (all have full stats: corners, shots, fouls, O/U odds):
  SC0  Scottish Premiership
  SC1  Scottish Championship
  SC2  Scottish League One
  SC3  Scottish League Two
  B1   Belgian First Division A
  N1   Dutch Eredivisie
  P1   Portuguese Primeira Liga
  P2   Portuguese Segunda Liga
  G1   Greek Super League
  T1   Turkish Süper Lig

Usage
-----
    python add_leagues.py            # download all leagues
    python add_leagues.py --dry-run  # test downloads without writing Excel/config
    python add_leagues.py --codes SC0 B1 N1   # subset of leagues
"""
from __future__ import annotations

import argparse
import logging
import sys
import time
from io import StringIO
from pathlib import Path

import numpy as np
import pandas as pd
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

# ── League definitions ────────────────────────────────────────────────────────

LEAGUES = {
    "PL1": {
        "name":         "Poland Ekstraklasa",
        "odds_api_key": "soccer_poland_ekstraklasa",
    },
    "SC0": {
        "name":         "Scottish Premiership",
        "odds_api_key": "soccer_scotland_premiership",
    },
    "SC1": {
        "name":         "Scottish Championship",
        "odds_api_key": None,
    },
    "SC2": {
        "name":         "Scottish League One",
        "odds_api_key": None,
    },
    "SC3": {
        "name":         "Scottish League Two",
        "odds_api_key": None,
    },
    "B1": {
        "name":         "Belgian First Division A",
        "odds_api_key": "soccer_belgium_first_div",
    },
    "N1": {
        "name":         "Dutch Eredivisie",
        "odds_api_key": "soccer_netherlands_eredivisie",
    },
    "P1": {
        "name":         "Portuguese Primeira Liga",
        "odds_api_key": "soccer_portugal_primeira_liga",
    },
    # P2 omitted — football-data.co.uk mmz4281/P2 returns Spanish Segunda data
    # (same as SP2/La Liga 2 which we already have). No clean source for Portuguese
    # Segunda Liga in this format.
    "G1": {
        "name":         "Greek Super League",
        "odds_api_key": "soccer_greece_super_league",
    },
    "T1": {
        "name":         "Turkish Super Lig",
        "odds_api_key": "soccer_turkey_super_league",
    },
}

# Internal season code → URL season segment used by mmz4281
# "2020" means 2019/20; on football-data.co.uk that folder is "1920"
SEASON_TO_URL = {
    "2020": "1920",
    "2021": "2021",
    "2122": "2122",
    "2223": "2223",
    "2324": "2324",
    "2425": "2425",
    "2526": "2526",
}

# Internal season code → human label (mirrors data_loader._SEASON_MAP)
SEASON_LABEL = {
    "2020": "2019/20",
    "2021": "2020/21",
    "2122": "2021/22",
    "2223": "2022/23",
    "2324": "2023/24",
    "2425": "2024/25",
    "2526": "2025/26",
}

BASE_URL = "https://www.football-data.co.uk/mmz4281/{season}/{code}.csv"

# Columns we want to pull through (all present in mmz4281 full-stats CSVs)
_COL_MAP = [
    ("home_goals",   "FTHG"),
    ("away_goals",   "FTAG"),
    ("home_corners", "HC"),
    ("away_corners", "AC"),
    ("home_fouls",   "HF"),
    ("away_fouls",   "AF"),
    ("home_shots",   "HS"),
    ("away_shots",   "AS"),
    ("home_sot",     "HST"),
    ("away_sot",     "AST"),
]

DATA_DIR = config.DATA_DIR / "data" / "football_data"


# ── Download ──────────────────────────────────────────────────────────────────

def _download_season(code: str, our_code: str) -> pd.DataFrame | None:
    url_season = SEASON_TO_URL[our_code]
    url = BASE_URL.format(season=url_season, code=code)
    try:
        r = requests.get(url, timeout=30)
        if r.status_code != 200:
            log.debug(f"  {code} {SEASON_LABEL[our_code]}: HTTP {r.status_code} — skipped")
            return None
        raw = pd.read_csv(
            StringIO(r.text),
            encoding="utf-8-sig",
            on_bad_lines="skip",
            low_memory=False,
        )
        # Must have at least Date + team columns + goals
        required = {"Date", "HomeTeam", "AwayTeam", "FTHG", "FTAG"}
        if not required.issubset(raw.columns):
            log.debug(f"  {code} {SEASON_LABEL[our_code]}: missing columns — skipped")
            return None
        raw = raw[raw["HomeTeam"].notna() & raw["AwayTeam"].notna()].copy()
        if raw.empty:
            return None
        return raw
    except Exception as e:
        log.warning(f"  {code} {SEASON_LABEL[our_code]}: {e}")
        return None


def download_league(code: str, name: str, dry_run: bool = False) -> dict[str, pd.DataFrame]:
    """Download all seasons for one league. Returns {our_code: raw_df}."""
    results = {}
    dest_dir = DATA_DIR / code

    for our_code in SEASON_TO_URL:
        raw = _download_season(code, our_code)
        if raw is None:
            continue

        log.info(f"  {name} {SEASON_LABEL[our_code]}: {len(raw)} rows")
        results[our_code] = raw

        if not dry_run:
            dest_dir.mkdir(parents=True, exist_ok=True)
            path = dest_dir / f"{code}_{our_code}.csv"
            raw.to_csv(path, index=False)

        time.sleep(0.3)   # be polite to football-data.co.uk

    return results


# ── Excel patcher ─────────────────────────────────────────────────────────────

def patch_excel(name: str, seasons: dict[str, pd.DataFrame]) -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.error("openpyxl not installed — run: pip install openpyxl")
        return 0

    wb = load_workbook(config.SUMMARY_XLSX)
    patched = 0

    for our_code, raw in seasons.items():
        sheet_name = f"{name}_{our_code}"

        # Replace if already exists, else create
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Write only the columns data_loader needs
        keep_cols = ["Date", "HomeTeam", "AwayTeam", "FTHG", "FTAG", "FTR",
                     "HC", "AC", "HF", "AF", "HS", "AS", "HST", "AST",
                     "Referee"]
        # Add whichever odds columns are present
        for odds_col in ["B365>2.5", "B365<2.5", "Avg>2.5", "Avg<2.5",
                         "Max>2.5", "Max<2.5", "AvgC>2.5", "AvgC<2.5"]:
            if odds_col in raw.columns:
                keep_cols.append(odds_col)
        keep_cols = [c for c in keep_cols if c in raw.columns]
        df_out = raw[keep_cols]

        for row in dataframe_to_rows(df_out, index=False, header=True):
            ws.append(row)

        log.info(f"  Patched sheet: {sheet_name} ({len(df_out)} rows)")
        patched += 1

    if patched:
        wb.save(config.SUMMARY_XLSX)
        log.info(f"Excel saved — {patched} new sheet(s) added")

    return patched


# ── Config updater ─────────────────────────────────────────────────────────────

def update_config(added: dict[str, str]) -> None:
    """
    added: {code: name} for all successfully downloaded leagues.
    Adds entries to FOOTBALL_DATA_LEAGUES and ODDS_API_SPORT_KEYS.
    Does NOT touch ENABLED_LEAGUES — run backtest first.
    """
    cfg_path = Path(__file__).resolve().parent / "config.py"
    text = cfg_path.read_text(encoding="utf-8")

    # ── FOOTBALL_DATA_LEAGUES ───────────────────────────────────────────────
    insert_fd = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        entry = f'    "{name}":{"" if len(name) >= 22 else " " * (22 - len(name))} "{code}",'
        if f'"{name}"' not in text:
            insert_fd += f"\n{entry}"

    if insert_fd:
        # Insert before the closing brace of FOOTBALL_DATA_LEAGUES
        marker = "}\n\nENABLED_LEAGUES"
        text = text.replace(marker, insert_fd + "\n" + marker, 1)
        log.info("Updated FOOTBALL_DATA_LEAGUES in config.py")

    # ── ODDS_API_SPORT_KEYS ─────────────────────────────────────────────────
    insert_odds = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        odds_key = LEAGUES[code].get("odds_api_key")
        if not odds_key:
            continue
        entry = f'    "{name}":{"" if len(name) >= 22 else " " * (22 - len(name))} "{odds_key}",'
        if f'"{name}"' not in text:
            insert_odds += f"\n{entry}"

    if insert_odds:
        # Insert before closing brace of ODDS_API_SPORT_KEYS
        marker = "}\n\nSOFASCORE_TOURNAMENT_IDS"
        text = text.replace(marker, insert_odds + "\n" + marker, 1)
        log.info("Updated ODDS_API_SPORT_KEYS in config.py")

    cfg_path.write_text(text, encoding="utf-8")


# ── data_loader CSV override updater ─────────────────────────────────────────

def update_data_loader(added: dict[str, str]) -> None:
    """
    Extend the CSV override block in data_loader.py to include new leagues,
    so if someone later downloads fresh per-season CSVs they take priority.
    """
    dl_path = Path(__file__).resolve().parent / "src" / "data_loader.py"
    text = dl_path.read_text(encoding="utf-8")

    # Find the existing override dict and extend it
    old_marker = '            "Denmark Superliga": "DNK",   # downloaded via download_league_data.py'
    if old_marker not in text:
        log.warning("Could not find CSV override block in data_loader.py — skipping")
        return

    new_entries = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        entry = f'            "{name}": "{code}",'
        if f'"{name}": "{code}"' not in text:
            new_entries += f"\n{entry}"

    if new_entries:
        text = text.replace(
            old_marker,
            old_marker + new_entries,
        )
        dl_path.write_text(text, encoding="utf-8")
        log.info(f"Updated CSV override map in data_loader.py ({len(added)} leagues)")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--codes", nargs="+", metavar="CODE",
                        help="Download only these FD codes (e.g. SC0 B1 N1)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Download and report but do not write Excel or config")
    args = parser.parse_args()

    targets = {k: v for k, v in LEAGUES.items()
               if not args.codes or k in args.codes}

    if not targets:
        log.error(f"No matching codes. Available: {list(LEAGUES.keys())}")
        sys.exit(1)

    log.info(f"Downloading {len(targets)} league(s): {list(targets.keys())}")
    log.info(f"Excel: {config.SUMMARY_XLSX}")
    if args.dry_run:
        log.info("DRY RUN — no files will be written")

    all_seasons: dict[str, dict[str, pd.DataFrame]] = {}

    for code, meta in targets.items():
        name = meta["name"]
        log.info(f"\n── {name} ({code}) ────────────────────────────")
        seasons = download_league(code, name, dry_run=args.dry_run)
        if seasons:
            all_seasons[code] = seasons
            log.info(f"  {name}: {len(seasons)} seasons downloaded")
        else:
            log.warning(f"  {name}: no data retrieved — skipped")

    if args.dry_run:
        print("\n" + "=" * 60)
        print("  DRY RUN RESULTS")
        print("=" * 60)
        for code, seasons in all_seasons.items():
            name = LEAGUES[code]["name"]
            total = sum(len(df) for df in seasons.values())
            print(f"  {name} ({code}): {len(seasons)} seasons, {total} rows total")
            for sc, df in sorted(seasons.items()):
                teams = df["HomeTeam"].dropna().unique()[:3].tolist()
                print(f"    {SEASON_LABEL[sc]}: {len(df)} rows | sample: {teams}")
        return

    # Write Excel + config
    log.info("\nPatching Excel workbook ...")
    for code, seasons in all_seasons.items():
        name = LEAGUES[code]["name"]
        patch_excel(name, seasons)

    added = {code: LEAGUES[code]["name"] for code in all_seasons}
    log.info("\nUpdating config.py ...")
    update_config(added)

    log.info("\nUpdating data_loader.py CSV override map ...")
    update_data_loader(added)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  RESULTS")
    print("=" * 60)
    for code, seasons in all_seasons.items():
        name = LEAGUES[code]["name"]
        total = sum(len(df) for df in seasons.values())
        print(f"\n  {name} ({code}): {len(seasons)} seasons, {total} rows")
        for sc, df in sorted(seasons.items()):
            teams = df["HomeTeam"].dropna().unique()[:3].tolist()
            print(f"    {SEASON_LABEL[sc]}: {len(df):>4} rows | {teams}")

    print()
    print("  Config updated. These leagues are NOT yet in ENABLED_LEAGUES.")
    print("  Run backtest to check profitability, then add winners manually:")
    print()
    print("    python backtest_all_leagues.py")
    print()
    print("  Then add profitable leagues to ENABLED_LEAGUES in config.py")
    print("  and retrain:")
    print()
    print("    python pipeline.py --mode all")


if __name__ == "__main__":
    main()
