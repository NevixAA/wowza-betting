"""
Download real league data from football-data.co.uk and patch the Excel workbook.

Usage
-----
    python download_league_data.py          # fix Denmark Superliga (default)
    python download_league_data.py --all    # all available leagues

What it does
------------
1. Downloads the DNK.csv "new-format" file (real Danish Superliga data)
2. Splits it by season into per-season CSVs in mixed/data/football_data/DNK/
3. Patches the Excel workbook — replaces contaminated Denmark sheets with real data
4. Re-enables Denmark Superliga in config.py

Notes
-----
- Denmark 1st Div has NO clean source on football-data.co.uk — stays disabled
- The "new" format has goals/results but NO over/under 2.5 odds or corner/shot stats
  → those features will be filled with league medians during training (acceptable)
"""
from __future__ import annotations

import argparse
import logging
import sys
import time
from io import StringIO
from pathlib import Path

import numpy as np
import pandas as pd
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

# ── Constants ─────────────────────────────────────────────────────────────────

# football-data.co.uk "new" format — single multi-season file per country
NEW_FORMAT_LEAGUES = {
    "Denmark Superliga": {
        "url":  "https://www.football-data.co.uk/new/DNK.csv",
        "code": "DNK",
        "filter_league": "Superliga",   # value in the 'League' column to keep
    },
}

# Maps DNK.csv "Season" column values to our internal season codes
_DNK_SEASON_MAP = {
    "2019/2020": "2020",
    "2020/2021": "2021",
    "2021/2022": "2122",
    "2022/2023": "2223",
    "2023/2024": "2324",
    "2024/2025": "2425",
    "2025/2026": "2526",
}
_SEASON_LABEL = {v: k.replace("/20", "/") for k, v in _DNK_SEASON_MAP.items()}
# e.g. "2020" → "2019/20"

DATA_DIR = config.DATA_DIR / "data" / "football_data"


# ── Download + parse ──────────────────────────────────────────────────────────

def _download_new_format(league_name: str, cfg: dict) -> dict[str, pd.DataFrame]:
    """
    Download and parse a football-data.co.uk "new" format file.
    Returns {our_season_code: standardised DataFrame}.
    """
    url = cfg["url"]
    log.info(f"Downloading {league_name} from {url} ...")

    r = requests.get(url, timeout=30)
    r.raise_for_status()

    raw = pd.read_csv(
        StringIO(r.text),
        encoding="utf-8-sig",
        on_bad_lines="skip",
        low_memory=False,
    )

    # Filter to the right competition (DNK.csv has only Superliga — but be safe)
    league_col = "League"
    if league_col in raw.columns and cfg.get("filter_league"):
        raw = raw[raw[league_col].str.strip() == cfg["filter_league"]].copy()

    results: dict[str, pd.DataFrame] = {}

    for dnk_season, our_code in _DNK_SEASON_MAP.items():
        chunk = raw[raw["Season"] == dnk_season].copy()
        if chunk.empty:
            continue

        # Standardise to the same column names the data_loader expects
        df = pd.DataFrame()
        df["Date"]     = chunk["Date"]
        df["HomeTeam"] = chunk["Home"]
        df["AwayTeam"] = chunk["Away"]
        df["FTHG"]     = pd.to_numeric(chunk["HG"], errors="coerce")
        df["FTAG"]     = pd.to_numeric(chunk["AG"], errors="coerce")
        df["FTR"]      = chunk.get("Res", np.nan)

        # No corners/shots/fouls/O-U odds in new format → leave as NaN
        # data_loader will fill these with league medians
        for col in ["HC","AC","HF","AF","HS","AS","HST","AST",
                    "B365>2.5","B365<2.5","Avg>2.5","Avg<2.5"]:
            df[col] = np.nan

        df = df[df["HomeTeam"].notna() & df["AwayTeam"].notna()].copy()

        dest_dir  = DATA_DIR / cfg["code"]
        dest_dir.mkdir(parents=True, exist_ok=True)
        path = dest_dir / f"{cfg['code']}_{our_code}.csv"
        df.to_csv(path, index=False)
        log.info(f"  {dnk_season}: {len(df)} rows → {path.name}")
        results[our_code] = df

    return results


# ── Excel patcher ─────────────────────────────────────────────────────────────

def patch_excel(league_name: str, seasons: dict[str, pd.DataFrame]) -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.error("openpyxl not installed — run: pip install openpyxl")
        return 0

    log.info(f"Patching Excel: {config.SUMMARY_XLSX}")
    wb = load_workbook(config.SUMMARY_XLSX)
    patched = 0

    for our_code, df in seasons.items():
        sheet_name = f"{league_name}_{our_code}"
        if sheet_name not in wb.sheetnames:
            log.warning(f"  Sheet not found: {sheet_name} — skipping")
            continue

        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        log.info(f"  ✓ Patched {sheet_name} ({len(df)} rows)")
        patched += 1

    if patched:
        wb.save(config.SUMMARY_XLSX)
        log.info(f"Excel saved — {patched} sheet(s) replaced")

    return patched


# ── Config updater ─────────────────────────────────────────────────────────────

def reenable_in_config(league_names: list[str]) -> None:
    cfg_path = Path(__file__).resolve().parent / "config.py"
    text = cfg_path.read_text(encoding="utf-8")
    changed = False
    for league in league_names:
        commented = f'    # "{league}"'
        uncommented = f'    "{league}"'
        if commented in text:
            text = text.replace(commented, uncommented)
            log.info(f"  Re-enabled {league} in config.py")
            changed = True
    # Remove the explanatory comment block if present
    block = (
        "    # Denmark leagues temporarily disabled — Excel sheets contain wrong data\n"
        "    # (Superliga sheets = Bundesliga 1, 1st Div sheets = Bundesliga 2 duplicate)\n"
        "    # Re-enable after running: python download_league_data.py\n"
    )
    if block in text:
        text = text.replace(block, "")
        changed = True
    if changed:
        cfg_path.write_text(text, encoding="utf-8")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true", help="Download all available leagues")
    args = parser.parse_args()

    targets = list(NEW_FORMAT_LEAGUES.keys()) if args.all else ["Denmark Superliga"]
    log.info(f"Targets: {targets}")

    fixed = []
    for league_name in targets:
        cfg = NEW_FORMAT_LEAGUES.get(league_name)
        if not cfg:
            log.warning(f"No download config for: {league_name}")
            continue
        seasons = _download_new_format(league_name, cfg)
        if seasons:
            patch_excel(league_name, seasons)
            fixed.append(league_name)

    if fixed:
        log.info(f"\nRe-enabling {fixed} in config.py ...")
        reenable_in_config(fixed)

    print("\n" + "=" * 60)
    print("  RESULT")
    print("=" * 60)
    for lg in fixed:
        cfg = NEW_FORMAT_LEAGUES[lg]
        code_dir = DATA_DIR / cfg["code"]
        files = sorted(code_dir.glob("*.csv"))
        print(f"\n  {lg} ({len(files)} season files):")
        for f in files:
            df = pd.read_csv(f)
            sample = df["HomeTeam"].dropna().head(3).tolist()
            print(f"    {f.name}: {len(df)} rows | sample: {sample}")

    print()
    if "Denmark 1st Div" not in fixed:
        print("  NOTE: Denmark 1st Div stays disabled — no clean data source found.")
        print("        football-data.co.uk does not publish Danish 1st Division data.")
    print()
    print("  Next step: python pipeline.py --mode all")


if __name__ == "__main__":
    main()
