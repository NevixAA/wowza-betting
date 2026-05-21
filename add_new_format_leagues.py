"""
Add low-market leagues from football-data.co.uk "new" format files.

These leagues have goals/results/1X2 odds but NO O/U odds or corners/shots.
The model trains on form features only for these leagues (sot_ratio unavailable,
implied_prob will be league median). At prediction time, OddsAPI provides live
O/U odds so edge IS computable for supported leagues.

Why add them anyway:
  - Bookmakers are less sharp on these markets → more exploitable edge
  - Form features (over25_last5, attack_str, defense_str) still work
  - Training data diversity improves overall model calibration

Leagues added:
  AUT   Austrian Bundesliga     (winter league, YYYY/YYYY seasons)
  ROM   Romanian Superliga      (winter league, YYYY/YYYY seasons)
  CHE   Swiss Super League      (winter league, single-year YYYY seasons)
  SWE   Sweden Allsvenskan      (summer league, single-year YYYY seasons)

Usage
-----
    python add_new_format_leagues.py
    python add_new_format_leagues.py --dry-run
"""
from __future__ import annotations

import argparse
import logging
import sys
import time
from io import StringIO
from pathlib import Path

import numpy as np
import pandas as pd
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

# ── League definitions ────────────────────────────────────────────────────────

LEAGUES = {
    "AUT": {
        "name":           "Austrian Bundesliga",
        "filter_league":  "Bundesliga",
        "season_type":    "winter",    # YYYY/YYYY format in file
        "odds_api_key":   "soccer_austria_bundesliga",
    },
    "ROM": {
        "name":           "Romanian Superliga",
        "filter_league":  "Superliga",
        "season_type":    "winter",
        "odds_api_key":   None,        # Not on OddsAPI — training-only
    },
    # CHE = China Super League on football-data.co.uk (NOT Switzerland) — excluded
    "SWE": {
        "name":           "Sweden Allsvenskan",
        "filter_league":  "Allsvenskan",
        "season_type":    "summer",
        "odds_api_key":   "soccer_sweden_allsvenskan",
    },
}

# Which seasons to keep (mirrors data_loader._SEASON_MAP)
# Winter leagues: "YYYY/YYYY" → internal code
WINTER_SEASON_MAP = {
    "2019/2020": "2020",
    "2020/2021": "2021",
    "2021/2022": "2122",
    "2022/2023": "2223",
    "2023/2024": "2324",
    "2024/2025": "2425",
    "2025/2026": "2526",
}
# Summer leagues: single "YYYY" → internal code
# (Sweden/Swiss run in a calendar year; we map to nearest winter code)
SUMMER_SEASON_MAP = {
    "2020": "2020",   # 2020 season → 2019/20
    "2021": "2021",   # 2021 season → 2020/21
    "2022": "2122",   # 2022 → 2021/22
    "2023": "2223",
    "2024": "2324",
    "2025": "2425",
    "2026": "2526",
}

SEASON_LABEL = {
    "2020": "2019/20", "2021": "2020/21", "2122": "2021/22",
    "2223": "2022/23", "2324": "2023/24", "2425": "2024/25",
    "2526": "2025/26",
}

BASE_URL = "https://www.football-data.co.uk/new/{code}.csv"


# ── Download + parse ──────────────────────────────────────────────────────────

def download_league(code: str, meta: dict) -> dict[str, pd.DataFrame]:
    """Download + parse a new-format file. Returns {our_code: standardised df}."""
    url = BASE_URL.format(code=code)
    log.info(f"Downloading {meta['name']} from {url} ...")
    r = requests.get(url, timeout=30)
    r.raise_for_status()

    raw = pd.read_csv(
        StringIO(r.text),
        encoding="utf-8-sig",
        on_bad_lines="skip",
        low_memory=False,
    )

    # Filter to the right competition
    filter_val = meta["filter_league"]
    if "League" in raw.columns:
        raw = raw[raw["League"].str.strip() == filter_val].copy()

    season_map = WINTER_SEASON_MAP if meta["season_type"] == "winter" else SUMMER_SEASON_MAP

    results = {}
    for file_season, our_code in season_map.items():
        chunk = raw[raw["Season"].astype(str).str.strip() == str(file_season)].copy()
        if chunk.empty:
            continue

        df = pd.DataFrame()
        df["Date"]     = chunk["Date"]
        df["HomeTeam"] = chunk["Home"]
        df["AwayTeam"] = chunk["Away"]
        df["FTHG"]     = pd.to_numeric(chunk["HG"], errors="coerce")
        df["FTAG"]     = pd.to_numeric(chunk["AG"], errors="coerce")
        df["FTR"]      = chunk.get("Res", np.nan)

        # No corners/shots/fouls/O-U odds in new format
        for col in ["HC", "AC", "HF", "AF", "HS", "AS", "HST", "AST",
                    "B365>2.5", "B365<2.5", "Avg>2.5", "Avg<2.5"]:
            df[col] = np.nan

        df = df[df["HomeTeam"].notna() & df["AwayTeam"].notna()].copy()
        if df.empty:
            continue

        log.info(f"  {meta['name']} {SEASON_LABEL[our_code]}: {len(df)} rows")
        results[our_code] = df

    return results


# ── Excel patcher ─────────────────────────────────────────────────────────────

def patch_excel(name: str, seasons: dict[str, pd.DataFrame]) -> int:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.error("openpyxl not installed — run: pip install openpyxl")
        return 0

    wb = load_workbook(config.SUMMARY_XLSX)
    patched = 0

    for our_code, df in seasons.items():
        sheet_name = f"{name}_{our_code}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        log.info(f"  Patched sheet: {sheet_name} ({len(df)} rows)")
        patched += 1

    if patched:
        wb.save(config.SUMMARY_XLSX)
        log.info(f"Excel saved — {patched} new sheet(s) added for {name}")

    return patched


# ── Config updater ─────────────────────────────────────────────────────────────

def update_config(added: dict[str, str]) -> None:
    cfg_path = Path(__file__).resolve().parent / "config.py"
    text = cfg_path.read_text(encoding="utf-8")

    insert_fd = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        entry = f'    "{name}": "{code}",'
        if f'"{name}"' not in text:
            insert_fd += f"\n{entry}"

    if insert_fd:
        marker = "}\n\nENABLED_LEAGUES"
        text = text.replace(marker, insert_fd + "\n" + marker, 1)
        log.info("Updated FOOTBALL_DATA_LEAGUES in config.py")

    insert_odds = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        odds_key = LEAGUES[code].get("odds_api_key")
        if not odds_key:
            continue
        entry = f'    "{name}": "{odds_key}",'
        if f'"{name}"' not in text:
            insert_odds += f"\n{entry}"

    if insert_odds:
        marker = "}\n\nSOFASCORE_TOURNAMENT_IDS"
        text = text.replace(marker, insert_odds + "\n" + marker, 1)
        log.info("Updated ODDS_API_SPORT_KEYS in config.py")

    cfg_path.write_text(text, encoding="utf-8")


# ── data_loader CSV override updater ─────────────────────────────────────────

def update_data_loader(added: dict[str, str]) -> None:
    dl_path = Path(__file__).resolve().parent / "src" / "data_loader.py"
    text = dl_path.read_text(encoding="utf-8")

    old_marker = '            "Denmark Superliga": "DNK",   # downloaded via download_league_data.py'
    if old_marker not in text:
        log.warning("Could not find CSV override block in data_loader.py — skipping")
        return

    new_entries = ""
    for code, name in sorted(added.items(), key=lambda x: x[1]):
        entry = f'            "{name}": "{code}",'
        if f'"{name}": "{code}"' not in text:
            new_entries += f"\n{entry}"

    if new_entries:
        text = text.replace(old_marker, old_marker + new_entries)
        dl_path.write_text(text, encoding="utf-8")
        log.info(f"Updated CSV override map in data_loader.py")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.dry_run:
        log.info("DRY RUN — no files will be written")

    all_seasons: dict[str, dict[str, pd.DataFrame]] = {}

    for code, meta in LEAGUES.items():
        log.info(f"\n── {meta['name']} ({code}) ────────────────────────────")
        try:
            seasons = download_league(code, meta)
        except Exception as e:
            log.error(f"  Failed: {e}")
            continue

        if seasons:
            all_seasons[code] = seasons
            log.info(f"  {meta['name']}: {len(seasons)} seasons, "
                     f"{sum(len(d) for d in seasons.values())} rows total")
        else:
            log.warning(f"  {meta['name']}: no usable data — skipped")

        time.sleep(0.5)

    if args.dry_run:
        print("\n=== DRY RUN RESULTS ===")
        for code, seasons in all_seasons.items():
            name = LEAGUES[code]["name"]
            total = sum(len(d) for d in seasons.values())
            print(f"  {name}: {len(seasons)} seasons, {total} rows")
        return

    log.info("\nPatching Excel workbook ...")
    for code, seasons in all_seasons.items():
        patch_excel(LEAGUES[code]["name"], seasons)

    added = {code: LEAGUES[code]["name"] for code in all_seasons}
    log.info("\nUpdating config.py ...")
    update_config(added)

    log.info("\nUpdating data_loader.py CSV override map ...")
    update_data_loader(added)

    print("\n" + "=" * 60)
    print("  RESULTS")
    print("=" * 60)
    for code, seasons in all_seasons.items():
        name = LEAGUES[code]["name"]
        total = sum(len(d) for d in seasons.values())
        odds_api = LEAGUES[code]["odds_api_key"]
        print(f"\n  {name} ({code}): {len(seasons)} seasons, {total} rows")
        print(f"    OddsAPI: {odds_api or 'NOT COVERED — training only'}")
        for sc, df in sorted(seasons.items()):
            teams = df["HomeTeam"].dropna().unique()[:3].tolist()
            print(f"    {SEASON_LABEL[sc]}: {len(df):>4} rows | {teams}")

    print()
    print("  These leagues are NOT yet in ENABLED_LEAGUES.")
    print("  Run backtest first, then enable the profitable ones.")
    print()
    print("    python backtest_all_leagues.py")


if __name__ == "__main__":
    main()
